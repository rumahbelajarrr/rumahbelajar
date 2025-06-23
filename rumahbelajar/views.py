from django.db import IntegrityError
from django.shortcuts import render, redirect
from .models import Absensi, QRCode, Siswa, Guru, JadwalLes, OrangTua, MataPelajaran, PresensiGuru, DetailPresensi, DailyQRCode
from .forms import AbsensiForm, PresensiGuruForm, ParentForm
from django.contrib.auth.decorators import login_required, user_passes_test
from django.shortcuts import render, redirect, get_object_or_404
import qrcode
from keuangan.models import Siswa as SiswaKeuangan
from django.http import HttpResponse
from io import BytesIO
from datetime import date, datetime
from django.contrib.auth import authenticate, login,logout
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from keuangan.models import PembayaranSPP, Siswa, OrangTua, Tagihan
from keuangan.forms import BuktiPembayaranForm
from .models import JadwalLes
from .models import Kelas
from django.http import JsonResponse
from .forms import GuruForm, SiswaForm, StudentForm
from django.http import HttpResponse
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from django.shortcuts import render, redirect
from django.utils import timezone
from rumahbelajar.models import MataPelajaran
from .models import PresensiSiswa
from .forms import PresensiForm
from django.contrib.auth.models import Group, User
from functools import wraps
import openpyxl
from django.db.models import Sum
import pytz
from django.core.exceptions import ValidationError
from .decorators import group_required
import qrcode
from django.core.files.base import ContentFile
import base64
from .utils import is_valid_time_hash, get_time_hash
import io
from django.contrib.auth import update_session_auth_hash
from django.urls import reverse
from django.db.models import Min, Max
from datetime import datetime, timedelta

def get_year_range_for_model(model, date_field):
    """Utility function to get year range for a model based on date field"""
    try:
        # Get the earliest and latest dates
        earliest = model.objects.aggregate(Min(date_field))[f'{date_field}__min']
        latest = model.objects.aggregate(Max(date_field))[f'{date_field}__max']
        
        if earliest and latest:
            start_year = earliest.year
            end_year = latest.year + 2  # Add 2 years to the latest year
        else:
            # If no data, use current year
            current_year = datetime.now().year
            start_year = current_year
            end_year = current_year + 2
        
        return list(range(start_year, end_year + 1))
    except Exception as e:
        # Fallback to current year if there's an error
        current_year = datetime.now().year
        return list(range(current_year, current_year + 3))

def group_required(group_name):
    def decorator(view_func):
        @wraps(view_func)
        def wrapper(request, *args, **kwargs):
            if request.user.groups.filter(name=group_name).exists():
                return view_func(request, *args, **kwargs)
            else:
                return redirect('rumahbelajar:login')
        return wrapper
    return decorator

@login_required
@group_required('Admin')
def presensi_guru_admin(request):
    if request.method == 'POST':
        form = PresensiGuruForm(request.POST)
        if form.is_valid():
            try:
                presensi = form.save(commit=False)
                # Pastikan guru diisi dari form
                guru_id = request.POST.get('guru')
                if guru_id:
                    presensi.guru_id = guru_id
                # Pastikan tanggal diisi dari form
                tanggal = request.POST.get('tanggal')
                if tanggal:
                    presensi.tanggal = tanggal
                # Pastikan jam_masuk diisi dari form
                jam_masuk = request.POST.get('jam_masuk')
                if jam_masuk:
                    presensi.jam_masuk = jam_masuk
                presensi.save()
                messages.success(request, 'Presensi guru berhasil disimpan')
                return redirect('rumahbelajar:presensi_guru_admin')
            except Exception as e:
                messages.error(request, f'Terjadi kesalahan: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = PresensiGuruForm()
    
    # Ambil data presensi berdasarkan filter
    presensi_guru_list = PresensiGuru.objects.all()
    tanggal = request.GET.get('tanggal')
    guru_id = request.GET.get('guru')
    status = request.GET.get('status')
    selected_year = request.GET.get('year', '')
    
    # Get year range for dropdown
    year_range = get_year_range_for_model(PresensiGuru, 'tanggal')
    
    if tanggal:
        presensi_guru_list = presensi_guru_list.filter(tanggal=tanggal)
    if guru_id:
        presensi_guru_list = presensi_guru_list.filter(guru_id=guru_id)
    if status:
        presensi_guru_list = presensi_guru_list.filter(status=status)
    if selected_year:
        presensi_guru_list = presensi_guru_list.filter(tanggal__year=selected_year)
    
    # Urutkan berdasarkan tanggal dan jam terbaru
    presensi_guru_list = presensi_guru_list.order_by('-tanggal', '-jam_masuk')
    
    # Ambil daftar guru untuk dropdown
    guru_list = Guru.objects.all().order_by('nama')
    
    context = {
        'form': form,
        'presensi_guru_list': presensi_guru_list,
        'guru_list': guru_list,
        'today': date.today(),
        'selected_year': selected_year,
        'year_range': year_range,
    }
    return render(request, 'admin/presensi_guru_admin.html', context)

@login_required
@group_required('Admin')
def export_presensi_guru_excel(request):
    # Buat workbook baru
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Presensi Guru"
    
    # Header
    headers = ['No', 'Nama Guru', 'Tanggal', 'Jam Masuk', 'Status', 'Keterangan']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Data
    presensi_list = PresensiGuru.objects.all().select_related('guru').order_by('-tanggal', '-jam_masuk')
    for row, presensi in enumerate(presensi_list, 2):
        ws.cell(row=row, column=1, value=row-1)  # Nomor urut
        ws.cell(row=row, column=2, value=presensi.guru.nama)
        ws.cell(row=row, column=3, value=presensi.tanggal.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=4, value=presensi.jam_masuk.strftime('%H:%M'))
        ws.cell(row=row, column=5, value=presensi.status)
        ws.cell(row=row, column=6, value=presensi.keterangan or '-')
    
    # Styling
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20
    
    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=presensi_guru.xlsx'
    wb.save(response)
    return response

@login_required
@group_required('Admin')
def export_presensi_guru_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="presensi_guru.pdf"'

    # Buat PDF
    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # Header
    p.setFont("Helvetica-Bold", 16)
    p.drawString(50, height - 50, "Laporan Presensi Guru")
    p.setFont("Helvetica", 10)
    p.drawString(50, height - 70, f"Tanggal Cetak: {date.today().strftime('%d/%m/%Y')}")
    
    # Subheader
    p.setFont("Helvetica-Bold", 10)
    headers = ['No', 'Nama Guru', 'Tanggal', 'Jam Masuk', 'Status', 'Keterangan']
    x_positions = [50, 150, 250, 350, 450, 500]
    y = height - 100
    
    for header, x in zip(headers, x_positions):
        p.drawString(x, y, header)
    
    # Data
    p.setFont("Helvetica", 10)
    y -= 20
    presensi_list = PresensiGuru.objects.all().select_related('guru').order_by('-tanggal', '-jam_masuk')
    
    for i, presensi in enumerate(presensi_list, 1):
        if y < 50:  # Jika halaman penuh, buat halaman baru
            p.showPage()
            y = height - 50
            p.setFont("Helvetica-Bold", 10)
            for header, x in zip(headers, x_positions):
                p.drawString(x, y, header)
            y -= 20
            p.setFont("Helvetica", 10)
        
        p.drawString(x_positions[0], y, str(i))
        p.drawString(x_positions[1], y, presensi.guru.nama)
        p.drawString(x_positions[2], y, presensi.tanggal.strftime('%d/%m/%Y'))
        p.drawString(x_positions[3], y, presensi.jam_masuk.strftime('%H:%M'))
        p.drawString(x_positions[4], y, presensi.status)
        p.drawString(x_positions[5], y, presensi.keterangan or '-')
        y -= 20

    p.showPage()
    p.save()
    return response
from django.utils.timezone import localdate
import datetime
@login_required
def absensi_siswa_view(request):
    try:
        siswa = Siswa.objects.get(user=request.user)
        kelas_siswa = siswa.kelas

        # Ambil semua mata pelajaran di kelas siswa
        mata_pelajaran_list = MataPelajaran.objects.filter(kelas=kelas_siswa).order_by('nama_pelajaran')

        context = {
            'mata_pelajaran_list': mata_pelajaran_list,
        }
        return render(request, 'siswa/absensi_daftar_mapel.html', context)
    except Siswa.DoesNotExist:
        messages.error(request, 'Akun siswa tidak valid.')
        return redirect('rumahbelajar:login')
@login_required
def absensi_detail(request, presensi_id):
    # Ambil jadwal presensi
    jadwal = get_object_or_404(Presensi, id=presensi_id)
    try:
        siswa = Siswa.objects.get(user=request.user)
        kelas_siswa = siswa.kelas

        # Untuk navigasi jika ingin ditampilkan daftar mapel siswa juga
        mata_pelajaran_list = MataPelajaran.objects.filter(kelas=kelas_siswa).order_by('nama_pelajaran')

        # Total pertemuan (contoh: 1 sampai 16)
        total_pertemuan = range(1, 17)

        # Ambil data presensi sebelumnya
        presensi_siswa = PresensiSiswa.objects.filter(presensi=jadwal, siswa=siswa)

        # Bentuk dict status kehadiran per pertemuan
        status_pertemuan = {
            p.pertemuan_ke: p.status for p in presensi_siswa if p.pertemuan_ke
        }

        # Bentuk dict waktu presensi per pertemuan
        presensi_waktu = {
            p.pertemuan_ke: p.waktu_presensi for p in presensi_siswa if p.pertemuan_ke
        }

        if request.method == "POST":
            pertemuan = request.POST.get('pertemuan')
            kode_qr = request.POST.get('kode_qr')

            # DEBUGGING
            print("====== DEBUG ABSENSI ======")
            print("JADWAL ID:", jadwal.id)
            print("PERTEMUAN:", pertemuan)
            print("KODE QR YANG DIKIRIM:", kode_qr)
            print("EXPECTED QR:", f"{jadwal.id}-{pertemuan}")
            print("===========================")

            # Validasi form
            if not pertemuan:
                messages.error(request, "Pertemuan harus diisi!")
                return redirect('rumahbelajar:absensi_detail', presensi_id=presensi_id)

            # Validasi QR code
            expected_qr = f"{jadwal.id}-{pertemuan}"
            if kode_qr.strip() != expected_qr.strip():
                messages.error(request, "Kode QR tidak valid.")
                return redirect('rumahbelajar:absensi_detail', presensi_id=presensi_id)

            # Simpan atau update presensi siswa
            presensi_obj, created = PresensiSiswa.objects.update_or_create(
                presensi=jadwal,
                siswa=siswa,
                pertemuan_ke=pertemuan,
                defaults={
                    'status': 'Hadir',
                    'waktu_presensi': timezone.now(),
                }
            )

            messages.success(request, f"Presensi pertemuan ke-{pertemuan} berhasil dicatat.")
            return redirect('rumahbelajar:absensi_detail', presensi_id=presensi_id)

        # Context untuk template
        context = {
            'jadwal': jadwal,
            'total_pertemuan': total_pertemuan,
            'status_pertemuan': status_pertemuan,
            'presensi_waktu': presensi_waktu,
            'mata_pelajaran_list': mata_pelajaran_list,
            'year': timezone.now().year,
        }

        return render(request, 'siswa/absensi_detail.html', context)
    except Siswa.DoesNotExist:
        messages.error(request, 'Akun siswa tidak valid.')
        return redirect('rumahbelajar:login')

@login_required
@group_required('Siswa')
def scan_qr(request, jadwal_id, pertemuan):
    if request.method == 'POST':
        kode_qr = request.POST.get('kode_qr')

        # QR: NamaUser-JadwalID-Pertemuan
        try:
            nama_user_qr, jadwal_id_qr, pertemuan_qr = kode_qr.strip().split('-')
            jadwal_id_qr = int(jadwal_id_qr)
            pertemuan_qr = int(pertemuan_qr)
        except ValueError:
            messages.error(request, "Format QR Code tidak valid.")
            return redirect('rumahbelajar:absensi_detail', jadwal_id=jadwal_id)

        if jadwal_id_qr != int(jadwal_id) or pertemuan_qr != int(pertemuan):
            messages.error(request, "QR Code tidak cocok dengan jadwal atau pertemuan.")
            return redirect('rumahbelajar:absensi_detail', jadwal_id=jadwal_id)

        try:
            siswa = Siswa.objects.get(user=request.user)
        except Siswa.DoesNotExist:
            messages.error(request, 'Akun siswa tidak valid.')
            return redirect('rumahbelajar:login')
            
        jadwal = get_object_or_404(MataPelajaran, id=jadwal_id)

        # Cek apakah sudah absen
        sudah_absen = Absensi.objects.filter(
            siswa=siswa,
            jadwal=jadwal,
            pertemuan=pertemuan
        ).exists()

        if sudah_absen:
            messages.warning(request, f"Kamu sudah absen untuk pertemuan {pertemuan}.")
        else:
            Absensi.objects.create(
                siswa=siswa,
                jadwal=jadwal,
                pertemuan=pertemuan,
                status='Hadir'  # Status default Hadir
            )
            messages.success(request, f"Absensi pertemuan {pertemuan} berhasil direkam.")

        return redirect('rumahbelajar:absensi_detail', jadwal_id=jadwal_id)

    
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import redirect
from django.utils.timezone import localdate, localtime
from .models import Absensi, Guru  # pastikan model Guru diimpor
import datetime
import json
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.utils.timezone import localdate
from django.contrib import messages
from .models import QRCode, PresensiGuru
from django.utils.timezone import now
import datetime
from django.shortcuts import render, redirect
from django.contrib import messages
from django.utils.timezone import now, localdate
from django.contrib.auth.decorators import login_required
import datetime
from .models import QRCode, PresensiGuru
from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect
from django.contrib import messages
from django.utils.timezone import localdate
import datetime
from .models import Absensi

from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect
from django.contrib import messages
from django.utils.timezone import localdate
import datetime
from .models import Absensi, Guru  # pastikan import model Guru
import logging
logger = logging.getLogger(__name__)
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.shortcuts import redirect
from django.utils.timezone import localdate
import datetime

@login_required
def generate_daily_qrcode(request, tanggal):
    print(f"DEBUG: generate_daily_qrcode dipanggil dengan tanggal: {tanggal}")
    print(f"DEBUG: User: {request.user.username}")
    
    # Validasi user harus guru
    try:
        guru_obj = Guru.objects.get(user=request.user)
        print(f"DEBUG: Guru ditemukan: {guru_obj.nama}")
    except Guru.DoesNotExist:
        print("DEBUG: User bukan guru")
        return JsonResponse({'status': 'error', 'message': 'Hanya guru yang bisa absen lewat QR code.'})

    # Validasi format tanggal
    try:
        tanggal_qr = datetime.datetime.strptime(tanggal, "%Y-%m-%d").date()
        print(f"DEBUG: Tanggal QR: {tanggal_qr}")
    except ValueError:
        print("DEBUG: Format tanggal tidak valid")
        return JsonResponse({'status': 'error', 'message': 'Format tanggal tidak valid.'})

    # Validasi QR hanya untuk hari ini
    tanggal_hari_ini = localdate()
    print(f"DEBUG: Tanggal hari ini: {tanggal_hari_ini}")
    if tanggal_qr != tanggal_hari_ini:
        print("DEBUG: QR bukan untuk hari ini")
        return JsonResponse({'status': 'error', 'message': 'QR code ini bukan untuk hari ini.'})

    # Simpan absensi hari ini ke database
    print("DEBUG: Mencoba menyimpan absensi...")
    absen, created = Absensi.objects.get_or_create(
        guru=guru_obj,
        tanggal=tanggal_hari_ini,
        defaults={
            'jam_absensi': timezone.localtime().time(),
            'status': 'Hadir',
            'metode_absensi': 'Absen via QR'
        }
    )
    
    if created:
        print(f"DEBUG: Absensi baru dibuat - ID: {absen.id}")
    else:
        print(f"DEBUG: Absensi sudah ada - ID: {absen.id}, update data...")
        absen.status = 'Hadir'
        absen.metode_absensi = 'Absen via QR'
        absen.jam_absensi = timezone.localtime().time()
        absen.save()
        print(f"DEBUG: Absensi diupdate - Status: {absen.status}, Metode: {absen.metode_absensi}")

    # Kembalikan response sukses dengan redirect URL
    response_data = {
        'status': 'success', 
        'message': 'Absensi berhasil dicatat!',
        'redirect_url': '/absen/presensi-guru/'
    }
    print(f"DEBUG: Mengirim response: {response_data}")
    return JsonResponse(response_data)

  # Redirect ke dashboard yang benar
@login_required
def scan_qr_view(request):
    today = timezone.localdate()
    print(today)
    context = {
        'today': today,# kirim ke template sebagai string "2025-05-28"
    }
    return render(request, 'guru/scan_guru.html',context)


@login_required
def jadwal_les(request):
    jadwals = JadwalLes.objects.all()
    return render(request, 'absen/jadwal_les.html', {'jadwals': jadwals})


@login_required
@group_required('Guru')
def absen_guru(request):
    if request.user.groups.filter(name='Guru').exists():
        guru = Guru.objects.get(user=request.user)
        if request.method == 'POST':
            form = AbsensiForm(request.POST)
            if form.is_valid():
                absensi = form.save(commit=False)
                absensi.guru = guru
                absensi.save()
                return redirect('absen_guru')  # Ganti dengan nama URL kamu
        else:
            form = AbsensiForm()
        return render(request, 'guru/absen_guru.html', {'form': form})
    else:
        return redirect('rumahbelajar:home')
    
@login_required
@group_required('Guru')
def absenkan_siswa(request, siswa_id):
    siswa = Siswa.objects.get(id=siswa_id)
    Absensi.objects.create(
        siswa=siswa,
        tanggal=date.today(),
        status='Hadir',
        metode_absensi='Input Manual'
    )
    return redirect('daftar_siswa.html')

@login_required
@group_required('Guru')
def lihat_absensi_siswa(request):
    absen = Absensi.objects.all().order_by('-tanggal')
    return render(request, 'siswa/absensi_siswa.html', {'absen': absen})


@login_required
@group_required('Guru')
def lihat_jadwal_guru(request):
    jadwal = JadwalLes.objects.all()
    return render(request, 'guru/jadwal.html', {'jadwal': jadwal})



@login_required
@group_required('Admin')
def dashboard_admin(request):
    semua_absen = Absensi.objects.all()
    semua_jadwal = JadwalLes.objects.all()
    return render(request, 'admin/dashboard.html', {
        'absen': semua_absen,
        'jadwal': semua_jadwal
    })
    
def redirect_dashboard(request):
    if request.user.groups.filter(name='Owner').exists():
        return redirect('rumahbelajar:dashboard_owner')
    if request.user.groups.filter(name='Admin').exists():
        return redirect('rumahbelajar:dashboard_admin')
    elif request.user.groups.filter(name='guru').exists():
        return redirect('rumahbelajar:dashboard_guru')
    elif request.user.groups.filter(name='Siswa').exists():
        return redirect('rumahbelajar:dashboard_Siswa')
    elif request.user.groups.filter(name='OrangTua').exists():
        return redirect('rumahbelajar:dashboard_OrangTua')
    else:
        return redirect('rumahbelajar:home')  # Jika role tidak terdaftar
 
def dashboard_guru(request):
    today = localdate()
    qr_code = DailyQRCode.objects.filter(tanggal=today).first()
    guru = Guru.objects.get(user=request.user)
    jadwal_list = JadwalLes.objects.filter(guru=guru)
    for j in jadwal_list:
        print(j.id, j.mata_pelajaran)
    context = {
        'jumlah_siswa': Siswa.objects.count(),
        'jumlah_kelas': Kelas.objects.count(),
        'jumlah_mapel': MataPelajaran.objects.count(),
        'today': today.strftime('%Y-%m-%d'),
        'qr_code': qr_code,
        'jadwal_list': jadwal_list
    }
    return render(request, 'guru/dashboard_guru.html', context)

@login_required
@group_required('Siswa')
def dashboard_siswa(request):
    return render(request, 'siswa/dashboard_siswa.html')


@login_required
def dashboard_orangtua(request):
    anak = get_object_or_404(Siswa, orangtua=request.user)
    total_kehadiran = Presensi.objects.filter(siswa=anak, status='Hadir').count()
    total_tagihan = Tagihan.objects.filter(siswa=anak).aggregate(Sum('jumlah'))['jumlah__sum'] or 0
    status_tagihan = 'Lunas' if not Tagihan.objects.filter(siswa=anak, status='Belum Lunas').exists() else 'Belum Lunas'

    context = {
        'anak': anak,
        'total_kehadiran': total_kehadiran,
        'total_tagihan': total_tagihan,
        'status_tagihan': status_tagihan,
    }
    return render(request, 'dashboard_orangtua.html', context)

def home(request):
    return render(request, 'absensi/home.html')

def logout_view(request):
    logout(request)
    return redirect('rumahbelajar:login')  # Menggunakan nama pattern 'login' yang terdaftar di urls.py

def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            if user.groups.filter(name='Owner').exists():
                return redirect('rumahbelajar:dashboard_owner')
            elif hasattr(user, 'Siswa'):
                return redirect('siswa:pembayaran_spp')
            return redirect('rumahbelajar:redirect_dashboard')  # Ganti ke halaman utama lo
        else:
            messages.error(request, 'Username atau password salah!')
    return render(request, 'absensi/login.html')  # Pastikan ada template login.html


@login_required
@group_required('OrangTua')
def dashboard_orangtua(request):
    try:
        orang_tua = request.user.rumahbelajar_orangtua  # Make sure the related name matches your model
        siswa_list = Siswa.objects.filter(orang_tua=orang_tua)
        tagihan_list = PembayaranSPP.objects.filter(siswa__in=siswa_list)

        context = {
            'orang_tua': orang_tua,
            'siswa': siswa_list,
            'tagihan': tagihan_list,
            'form': BuktiPembayaranForm(),  # if you're using the upload form
        }

        return render(request, 'orangtua/dashboard_orangtua.html', context)

    except OrangTua.DoesNotExist:
        return redirect('rumahbelajar:home')  # or handle differently


def jadwal_les(request):
    jadwals = JadwalLes.objects.all()  # Replace with actual model query
    return render(request, 'jadwal_les.html', {'jadwals': jadwals})

# rumahbelajar/views.py
from django.shortcuts import render

def daftar_siswa(request):
    siswa_list = Siswa.objects.all()  # Mengambil semua data siswa dari database
    return render(request, 'guru/daftar_siswa.html', {'siswa_list': siswa_list})

def akademik(request):
    # Ambil data kelas dan mata pelajaran dari database
    kelas_list = Kelas.objects.all()  # atau sesuaikan dengan query yang sesuai
    
    # Render halaman akademik dengan data kelas dan mata pelajaran
    return render(request, 'guru/akademik.html', {'kelas_list': kelas_list})

@login_required
def akademik_siswa_view(request):
    try:
        siswa = Siswa.objects.get(user=request.user)
        kelas = siswa.kelas
        mata_pelajaran_list = MataPelajaran.objects.filter(kelas=kelas) if kelas else []

        context = {
            'siswa': siswa,
            'kelas': kelas,
            'mata_pelajaran_list': mata_pelajaran_list,
        }

        return render(request, 'siswa/akademik_siswa.html', context)
    except Siswa.DoesNotExist:
        messages.error(request, 'Akun siswa tidak valid.')
        return redirect('rumahbelajar:login')


def dashboard_admin(request):
    jumlah_siswa = Siswa.objects.count()
    jumlah_kelas = Kelas.objects.count()
    jumlah_mapel = MataPelajaran.objects.count()
    jumlah_guru = Guru.objects.count()
    
    context = {
        'jumlah_siswa': jumlah_siswa,
        'jumlah_kelas': jumlah_kelas,
        'jumlah_mapel': jumlah_mapel,
        'jumlah_guru': jumlah_guru,
    }
    return render(request, 'admin/dashboard_admin.html', context)
from django.shortcuts import render

def contact(request):
    return render(request, 'absensi/contact.html')

def data_siswa(request):
    # Check if deletion is requested
    if 'delete' in request.GET:
        student_id = request.GET.get('delete')
        try:
            student = Siswa.objects.get(id=student_id)
            if student.user:
                student.user.delete()
            messages.success(request, f'Siswa {student.nama} berhasil dihapus')
        except Siswa.DoesNotExist:
            messages.error(request, 'Siswa tidak ditemukan')
        return redirect('rumahbelajar:data_siswa')

    # Ambil semua siswa dengan relasi orang tua dan kelas
    students = Siswa.objects.select_related('orang_tua', 'kelas').all()
    
    if request.method == 'POST':
        siswa_id = request.POST.get('siswa_id')
        if siswa_id:  # Edit siswa
            siswa = get_object_or_404(Siswa, id=siswa_id)
            form = StudentForm(request.POST, instance=siswa)
            if form.is_valid():
                form.save()
                messages.success(request, f'Data siswa {siswa.nama} berhasil diperbarui.')
                return redirect('rumahbelajar:data_siswa')
            else:
                messages.error(request, 'Gagal memperbarui data siswa. Periksa kembali formulir.')
        else:  # Tambah siswa baru
            form = StudentForm(request.POST)
            if form.is_valid():
                student = form.save()
                messages.success(request, f'Siswa {student.nama} berhasil ditambahkan')
                return redirect('rumahbelajar:data_siswa')

            else:
                messages.error(request, 'Gagal menambahkan siswa. Periksa kembali formulir.')
    else:
        form = StudentForm()

    orang_tua_list = OrangTua.objects.all()
    kelas_list = Kelas.objects.all()
        
    context = {
        'siswa_list': students,
        'form': form,
        'orang_tua_list': orang_tua_list,
        'kelas_list': kelas_list
    }
    
    return render(request, 'admin/data_siswa.html', context)


@login_required
@group_required('Admin')
def data_orang_tua(request):
    # Handle hapus
    if 'delete' in request.GET:
        parent_id = request.GET.get('delete')
        try:
            parent = OrangTua.objects.get(id=parent_id)
            if parent.user:
                parent.user.delete()
            messages.success(request, f'Data orang tua {parent.nama} berhasil dihapus')
        except OrangTua.DoesNotExist:
            messages.error(request, 'Data orang tua tidak ditemukan')
        return redirect('rumahbelajar:data_orang_tua')

    # Ambil semua data orang tua
    parents = OrangTua.objects.all()

    # Cek apakah sedang edit
    edit_id = request.GET.get('edit')
    instance = None
    if edit_id:
        instance = get_object_or_404(OrangTua, id=edit_id)

    if request.method == 'POST':
        # Cek apakah ini edit atau tambah baru
        if request.POST.get('parent_id'):
            instance = get_object_or_404(OrangTua, id=request.POST.get('parent_id'))
            form = ParentForm(request.POST, instance=instance)
            if form.is_valid():
                form.save()
                messages.success(request, f'Data orang tua {instance.nama} berhasil diperbarui')
                return redirect('rumahbelajar:data_orang_tua')
            else:
                messages.error(request, 'Terjadi kesalahan saat mengedit data.')
        else:
            form = ParentForm(request.POST)
            if form.is_valid():
                new_parent = form.save()
                messages.success(request, f'Data orang tua {new_parent.nama} berhasil ditambahkan')
                return redirect('rumahbelajar:data_orang_tua')
            else:
                messages.error(request, 'Terjadi kesalahan saat menambahkan data.')
    else:
        form = ParentForm(instance=instance)

    context = {
        'orang_tua_list': parents,
        'form': form,
        'edit_mode': bool(instance),
        'editing_id': instance.id if instance else None
    }

    return render(request, 'admin/data_orang_tua.html', context)

@login_required
@group_required('Admin')
def get_parent_data(request, parent_id):
    """API endpoint to get parent data for editing"""
    try:
        parent = OrangTua.objects.get(id=parent_id)
        data = {
            'id': parent.id,
            'nama': parent.nama,
            'alamat': parent.alamat,
            'no_telp': parent.no_telp,
        }
        return JsonResponse({'status': 'success', 'data': data})
    except OrangTua.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Data orang tua tidak ditemukan'})

@login_required
@group_required('Admin')
def update_parent(request, parent_id):
    """API endpoint to update parent data"""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'})
    
    try:
        parent = OrangTua.objects.get(id=parent_id)
        
        # Get data from POST
        nama = request.POST.get('nama')
        alamat = request.POST.get('alamat')
        no_telp = request.POST.get('no_telp')
        
        # Update parent data
        parent.nama = nama
        parent.alamat = alamat
        parent.no_telp = no_telp
        parent.save()
        
        messages.success(request, f'Data orang tua {nama} berhasil diupdate')
        return redirect('rumahbelajar:data_orang_tua')
    except Exception as e:
        messages.error(request, f'Error: {str(e)}')
        return redirect('rumahbelajar:data_orang_tua')
    
def data_guru_admin(request):
    # Handle penghapusan
    hapus_id = request.GET.get('hapus_id')
    if hapus_id:
        guru_hapus = get_object_or_404(Guru, id=hapus_id)
        user_hapus = guru_hapus.user  # Ambil user yang terkait
        guru_hapus.delete()           # Hapus Guru
        user_hapus.delete()           # Hapus User juga
        messages.success(request, 'Data guru dan user berhasil dihapus.')
        return redirect('rumahbelajar:data_guru_admin')

    # Handle edit
    edit_id = request.GET.get('edit_id')
    if edit_id:
        guru_instance = get_object_or_404(Guru, id=edit_id)
    else:
        guru_instance = None

    if request.method == 'POST':
        if 'edit_id' in request.POST:
            guru_instance = get_object_or_404(Guru, id=request.POST.get('edit_id'))
            # Karena kita tidak ingin edit username/password user, gunakan ModelForm biasa
            form = GuruForm(request.POST, instance=guru_instance)
            if form.is_valid():
                form.save()
                messages.success(request, 'Data guru berhasil diperbarui.')
                return redirect('rumahbelajar:data_guru_admin')
            else:
                messages.error(request, 'Gagal memperbarui data guru.')
        else:
            form = GuruForm(request.POST)
            if form.is_valid():
                form.save()  # save() sudah buat user & assign ke grup 'guru'
                messages.success(request, 'Data guru dan user berhasil ditambahkan.')
                return redirect('rumahbelajar:data_guru_admin')
            else:
                messages.error(request, 'Gagal menambahkan data guru.')
    else:
        form = GuruForm(instance=guru_instance)

    guru_list = Guru.objects.all()
    return render(request, 'admin/data_guru_admin.html', {
        'form': form,
        'guru_list': guru_list,
        'edit_id': edit_id
    })

from .models import PresensiSiswa, Kelas  # Pastikan model Kelas di-import
@login_required
@group_required('Admin')
def presensi_siswa_admin(request):
    today = timezone.now().date()

    # Ambil filter dari request
    selected_date = request.GET.get('tanggal', today.strftime('%Y-%m-%d'))
    selected_kelas = request.GET.get('kelas', '')
    search_nama = request.GET.get('nama', '')
    selected_year = request.GET.get('year', '')
    
    # Get year range for dropdown
    year_range = get_year_range_for_model(PresensiSiswa, 'waktu_presensi')

    # Ambil data presensi
    presensi_siswa_list = PresensiSiswa.objects.select_related(
        'siswa__kelas', 'presensi__mata_pelajaran'
    ).filter(
        waktu_presensi__date=selected_date
    )

    # Apply year filter if provided
    if selected_year:
        presensi_siswa_list = presensi_siswa_list.filter(waktu_presensi__year=selected_year)

    if selected_kelas:
        presensi_siswa_list = presensi_siswa_list.filter(siswa__kelas__nama=selected_kelas)

    if search_nama:
        presensi_siswa_list = presensi_siswa_list.filter(siswa__nama__icontains=search_nama)

    # Ambil semua nama kelas unik
    kelas_list = Siswa.objects.values_list('kelas__nama', flat=True).distinct()

    context = {
        'presensi_siswa_list': presensi_siswa_list,
        'today': selected_date,
        'kelas_list': kelas_list,
        'selected_kelas': selected_kelas,
        'search_nama': search_nama,
        'selected_year': selected_year,
        'year_range': year_range,
    }
    return render(request, 'admin/presensi_siswa_admin.html', context)


def rekap_keuangan_admin(request):
    return render(request, 'admin/rekap_keuangan_admin.html')

def kelola_pembayaran_admin(request):
    return render(request, 'admin/kelola_pembayaran_admin.html')

from django.urls import reverse

from django.shortcuts import render, get_object_or_404
from .models import Presensi, Siswa
def presensi_siswa(request):
    # Validasi guru harus sudah absen hari ini sebelum bisa melihat presensi siswa
    if request.user.groups.filter(name='Guru').exists():
        try:
            guru = Guru.objects.get(user=request.user)
            today = timezone.now().date()
            guru_has_attended = PresensiGuru.objects.filter(
                guru=guru,
                tanggal=today,
                status='Hadir'
            ).exists()
            
            if not guru_has_attended:
                messages.warning(request, 'Anda harus melakukan presensi terlebih dahulu sebelum dapat melihat presensi siswa!')
                return redirect('rumahbelajar:presensi_guru')
        except Guru.DoesNotExist:
            messages.error(request, 'Data guru tidak ditemukan!')
            return redirect('rumahbelajar:dashboard_guru')

    # Ambil semua presensi beserta daftar siswa (PresensiSiswa) yang hadir
    presensi_siswa_list = Presensi.objects.prefetch_related('presensi_siswa__siswa', 'mata_pelajaran', 'kelas')
    kelas_list = Kelas.objects.all()
    context = {
        'presensi_siswa_list': presensi_siswa_list,
        'kelas_list' : kelas_list
    }
    return render(request, 'guru/presensi_siswa.html', context)

from .models import Presensi, PresensiSiswa
from .models import Siswa, PresensiSiswa, Presensi
def detail_presensi(request, kelas_id):
    kelas = get_object_or_404(Kelas, id=kelas_id)
    pertemuan_dipilih = request.GET.get('pertemuan')

    try:
        pertemuan_dipilih = int(pertemuan_dipilih) if pertemuan_dipilih else None
    except ValueError:
        pertemuan_dipilih = None

    presensi_objek = Presensi.objects.filter(kelas=kelas).first()
    siswa_kelas = Siswa.objects.filter(kelas=kelas).select_related('kelas')

    # Validasi guru harus sudah absen hari ini sebelum bisa mencatat presensi siswa
    if request.user.groups.filter(name='Guru').exists():
        try:
            guru = Guru.objects.get(user=request.user)
            today = timezone.now().date()
            guru_has_attended = PresensiGuru.objects.filter(
                guru=guru,
                tanggal=today,
                status='Hadir'
            ).exists()
            
            if not guru_has_attended:
                messages.warning(request, 'Anda harus melakukan presensi terlebih dahulu sebelum dapat mencatat presensi siswa!')
                return redirect('rumahbelajar:presensi_guru')
        except Guru.DoesNotExist:
            messages.error(request, 'Data guru tidak ditemukan!')
            return redirect('rumahbelajar:dashboard_guru')

    data_presensi = []

    for siswa in siswa_kelas:
        presensi = PresensiSiswa.objects.filter(
            siswa=siswa,
            presensi__kelas=kelas,
            pertemuan_ke=pertemuan_dipilih
        ).select_related('presensi').first() if pertemuan_dipilih else None

        data_presensi.append({
            'siswa': siswa,
            'status': presensi.status if presensi else 'Belum Diisi',
            'presensi_id': presensi.id if presensi else None,
            'jam': presensi.waktu_presensi if presensi else None,
        })

    # Hitung statistik
    total_hadir = sum(1 for p in data_presensi if p['status'] == 'Hadir')
    total_tidak_hadir = sum(1 for p in data_presensi if p['status'] == 'Tidak Hadir')
    total_siswa = siswa_kelas.count()

    if request.method == 'POST' and pertemuan_dipilih:
        # Validasi lagi sebelum menyimpan (double check)
        if request.user.groups.filter(name='Guru').exists():
            try:
                guru = Guru.objects.get(user=request.user)
                today = timezone.now().date()
                guru_has_attended = PresensiGuru.objects.filter(
                    guru=guru,
                    tanggal=today,
                    status='Hadir'
                ).exists()
                
                if not guru_has_attended:
                    messages.error(request, 'Anda harus melakukan presensi terlebih dahulu sebelum dapat mencatat presensi siswa!')
                    return redirect('rumahbelajar:presensi_guru')
            except Guru.DoesNotExist:
                messages.error(request, 'Data guru tidak ditemukan!')
                return redirect('rumahbelajar:dashboard_guru')

        for item in data_presensi:
            siswa = item['siswa']
            status_input = request.POST.get(f'status_{siswa.id}')
            if status_input:
                presensi_obj, created = PresensiSiswa.objects.get_or_create(
                    siswa=siswa,
                    pertemuan_ke=pertemuan_dipilih,
                    presensi=presensi_objek,
                    defaults={'status': status_input}
                )
                if not created:
                    presensi_obj.status = status_input
                    presensi_obj.save()
        
        messages.success(request, f'Presensi pertemuan ke-{pertemuan_dipilih} berhasil disimpan!')
        return redirect(f"{reverse('rumahbelajar:detail_presensi', args=[kelas_id])}?pertemuan={pertemuan_dipilih}")

    return render(request, 'guru/detail_presensi.html', {
        'kelas': kelas,
        'data_presensi': data_presensi,
        'total_hadir': total_hadir,
        'total_tidak_hadir': total_tidak_hadir,
        'total_siswa': total_siswa,
        'pertemuan_dipilih': pertemuan_dipilih,
        'presensi': presensi_objek,
        'pertemuan_range': range(1, 17),
    })


def ubah_status_presensi(request, presensi_id):
    if request.method == 'POST':
        # Validasi guru harus sudah absen hari ini sebelum bisa mengubah status presensi siswa
        if request.user.groups.filter(name='Guru').exists():
            try:
                guru = Guru.objects.get(user=request.user)
                today = timezone.now().date()
                guru_has_attended = PresensiGuru.objects.filter(
                    guru=guru,
                    tanggal=today,
                    status='Hadir'
                ).exists()
                
                if not guru_has_attended:
                    messages.error(request, 'Anda harus melakukan presensi terlebih dahulu sebelum dapat mengubah status presensi siswa!')
                    return redirect('rumahbelajar:presensi_guru')
            except Guru.DoesNotExist:
                messages.error(request, 'Data guru tidak ditemukan!')
                return redirect('rumahbelajar:dashboard_guru')

        status = request.POST.get('status')
        presensi_item = get_object_or_404(PresensiSiswa, id=presensi_id)
        presensi_item.status = status
        presensi_item.save()
        messages.success(request, 'Status presensi berhasil diubah!')
    return redirect('rumahbelajar:presensi_detail', presensi_item.presensi.id)



def simpan_presensi(request, presensi_id):
    if request.method == 'POST':
        # Validasi guru harus sudah absen hari ini sebelum bisa menyimpan presensi siswa
        if request.user.groups.filter(name='Guru').exists():
            try:
                guru = Guru.objects.get(user=request.user)
                today = timezone.now().date()
                guru_has_attended = PresensiGuru.objects.filter(
                    guru=guru,
                    tanggal=today,
                    status='Hadir'
                ).exists()
                
                if not guru_has_attended:
                    messages.error(request, 'Anda harus melakukan presensi terlebih dahulu sebelum dapat menyimpan presensi siswa!')
                    return redirect('rumahbelajar:presensi_guru')
            except Guru.DoesNotExist:
                messages.error(request, 'Data guru tidak ditemukan!')
                return redirect('rumahbelajar:dashboard_guru')

        presensi = get_object_or_404(Presensi, id=presensi_id)
        siswa_list = Siswa.objects.filter(kelas=presensi.kelas)

        for siswa in siswa_list:
            status = request.POST.get(f'status_{siswa.id}')
            if status:
                DetailPresensi.objects.create(
                    presensi=presensi,
                    siswa=siswa,
                    status=status
                )
        messages.success(request, 'Presensi berhasil disimpan.')
        return redirect('rumahbelajar:detail_presensi')


from django.http import JsonResponse

@login_required
@group_required('Admin')
def get_student_data(request, student_id):
    student = get_object_or_404(Siswa, id=student_id)
    data = {
        'id': student.id,
        'username': student.user.username,
        'nama': student.nama,
        'tanggal_lahir': student.tanggal_lahir.strftime('%Y-%m-%d'),
        'jenis_kelamin': student.jenis_kelamin,
        'orang_tua_id': student.orang_tua.id if student.orang_tua else None,
        'kelas_id': student.kelas.id if student.kelas else None,
    }
    return JsonResponse({'status': 'success', 'data': data})

@login_required
@group_required('Admin')
def kelas(request):
    if request.method == 'POST':
        kelas_id = request.POST.get('id')
        nama = request.POST.get('nama')
        
        if kelas_id:  # Update existing kelas
            kelas = get_object_or_404(Kelas, id=kelas_id)
            kelas.nama = nama
            kelas.save()
            messages.success(request, 'Data kelas berhasil diperbarui')
        else:  # Create new kelas
            Kelas.objects.create(nama=nama)
            messages.success(request, 'Data kelas berhasil ditambahkan')
        
        return redirect('rumahbelajar:kelas')
    
    # Handle delete request
    delete_id = request.GET.get('delete')
    if delete_id:
        kelas = get_object_or_404(Kelas, id=delete_id)
        kelas.delete()
        messages.success(request, 'Data kelas berhasil dihapus')
        return redirect('rumahbelajar:kelas')

    # Handle edit request
    edit_id = request.GET.get('edit')
    kelas_edit = None
    if edit_id:
        kelas_edit = get_object_or_404(Kelas, id=edit_id)
    
    # Get all classes
    kelas_list = Kelas.objects.all().order_by('nama')
    
    context = {
        'kelas_list': kelas_list,
        'kelas_edit': kelas_edit,
    }
    return render(request, 'admin/kelas.html', context)


@login_required
@group_required('Admin')
def kelas_detail(request, kelas_id):
    kelas = get_object_or_404(Kelas, id=kelas_id)
    siswa_list = Siswa.objects.filter(kelas=kelas).order_by('nama')

    context = {
        'kelas': kelas,
        'siswa_list': siswa_list,
    }
    return render(request, 'admin/kelas_detail.html', context)


@login_required
@group_required('Admin')
def get_kelas_data(request, kelas_id):
    kelas = get_object_or_404(Kelas, id=kelas_id)
    data = {
        'id': kelas.id,
        'nama': kelas.nama,
    }
    return JsonResponse({'status': 'success', 'data': data})

@login_required
@group_required('Admin')
def edit_kelas(request, kelas_id):
    kelas = get_object_or_404(Kelas, id=kelas_id)
    if request.method == 'POST':
        nama_baru = request.POST.get('nama')
        if nama_baru:
            kelas.nama = nama_baru
            kelas.save()
    return redirect('rumahbelajar:kelas')

@login_required
@group_required('Admin')
def update_student(request, student_id):
    """API endpoint to update student data"""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Method not allowed'})
    
    try:
        student = Siswa.objects.get(id=student_id)
        
        # Get data from POST
        nama = request.POST.get('nama')
        tanggal_lahir = request.POST.get('tanggal_lahir')
        jenis_kelamin = request.POST.get('jenis_kelamin')
        orang_tua_id = request.POST.get('orang_tua')
        
        # Update student data
        student.nama = nama
        student.tanggal_lahir = tanggal_lahir
        student.jenis_kelamin = jenis_kelamin
        
        if orang_tua_id:
            from .models import OrangTua
            student.orang_tua = OrangTua.objects.get(id=orang_tua_id)
        
        student.save()
        messages.success(request, f'Data siswa {nama} berhasil diupdate')
        return redirect('rumahbelajar:data_siswa')
    except Exception as e:
        messages.error(request, f'Error: {str(e)}')
        return redirect('rumahbelajar:data_siswa')

# views.py
from django.shortcuts import render, get_object_or_404, redirect
from .models import Presensi, PresensiSiswa
from .forms import PresensiForm

# Fungsi untuk menambah presensi
def tambah_presensi(request, presensi_id):
    # Ambil objek Presensi berdasarkan presensi_id
    presensi = get_object_or_404(Presensi, id=presensi_id)

    # Jika requestnya adalah POST (form dikirim)
    if request.method == 'POST':
        form = PresensiForm(request.POST)
        if form.is_valid():
            # Simpan presensi siswa ke database
            presensi_siswa = form.save(commit=False)
            presensi_siswa.presensi = presensi  # Hubungkan dengan presensi yang sudah ada
            presensi_siswa.save()
            # Redirect ke halaman lain setelah sukses
            return redirect('some_view_name')  # Ganti dengan nama URL tujuan setelah berhasil
    else:
        form = PresensiForm()  # Form kosong untuk pertama kali

    # Render template dengan form dan objek presensi
    return render(request, 'rumahbelajar/tambah_presensi.html', {'form': form, 'presensi': presensi})


@login_required
def presensi_guru(request):
    try:
        guru = Guru.objects.get(user=request.user)
    except Guru.DoesNotExist:
        # Buat data guru baru jika belum ada
        guru = Guru.objects.create(
            user=request.user,
            nama=request.user.username,
            nip='',
            alamat_guru=''
        )
        messages.info(request, 'Data guru Anda telah dibuat. Silakan lengkapi profil Anda.')

    # Gunakan timezone Asia/Jakarta
    jakarta_tz = pytz.timezone('Asia/Jakarta')
    now = timezone.now().astimezone(jakarta_tz)
    today = now.date()
    
    has_submitted_today = PresensiGuru.objects.filter(
        guru=guru,
        tanggal=today
    ).exists()

    if request.method == 'POST':
        form = PresensiGuruForm(request.POST)
        if form.is_valid():
            presensi = form.save(commit=False)
            presensi.guru = guru
            presensi.tanggal = today
            presensi.jam_masuk = now.time()
            presensi.save()
            messages.success(request, 'Presensi berhasil disimpan!')
            return redirect('rumahbelajar:presensi_guru')
    else:
        form = PresensiGuruForm()

    presensi_list = PresensiGuru.objects.filter(guru=guru).order_by('-tanggal', '-jam_masuk')

    context = {
        'form': form,
        'presensi_list': presensi_list,
        'has_submitted_today': has_submitted_today,
        'today': today,
        'current_time': now.strftime('%H:%M:%S'),
    }
    return render(request, 'guru/presensi_guru.html', context)

@login_required
def daftar_presensi(request):
    guru_login = request.user.rumahbelajar_guru
    presensi_list = PresensiGuru.objects.filter(guru=guru_login).order_by('-tanggal', '-jam_masuk')
    return render(request, 'presensi/daftar_presensi.html', {'presensi_list': presensi_list})

@login_required
@group_required('Owner')
def dashboard_owner(request):
    jumlah_siswa = Siswa.objects.count()
    jumlah_kelas = Kelas.objects.count()
    jumlah_guru = Guru.objects.count()
    total_keuangan = PembayaranSPP.objects.aggregate(Sum('jumlah_bayar'))['jumlah_bayar__sum'] or 0
    context = {
        'jumlah_siswa': jumlah_siswa,
        'jumlah_kelas': jumlah_kelas,
        'jumlah_guru': jumlah_guru,
        'total_keuangan': total_keuangan,
    }
    return render(request, 'owner/dashboard_owner.html', context)


@login_required
def mata_pelajaran_admin(request):
    if request.method == 'POST':
        mapel_id = request.POST.get('id')
        nama_pelajaran = request.POST.get('nama_pelajaran')
        kode_pelajaran = request.POST.get('kode_pelajaran')
        kelas_id = request.POST.get('kelas')

        if mapel_id:  # Update existing mata pelajaran
            mapel = get_object_or_404(MataPelajaran, id=mapel_id)
            mapel.nama_pelajaran = nama_pelajaran
            mapel.kode_pelajaran = kode_pelajaran
            mapel.kelas_id = kelas_id
            mapel.save()
            messages.success(request, 'Mata pelajaran berhasil diperbarui!')
        else:  # Create new mata pelajaran
            MataPelajaran.objects.create(
                nama_pelajaran=nama_pelajaran,
                kode_pelajaran=kode_pelajaran,
                kelas_id=kelas_id
            )
            messages.success(request, 'Mata pelajaran berhasil ditambahkan!')
        return redirect('rumahbelajar:mata_pelajaran_admin')

    # Handle delete request
    delete_id = request.GET.get('delete')
    if delete_id:
        mapel = get_object_or_404(MataPelajaran, id=delete_id)
        mapel.delete()
        messages.success(request, 'Mata pelajaran berhasil dihapus!')
        return redirect('rumahbelajar:mata_pelajaran_admin')

    # Get all mata pelajaran and kelas for the template
    mata_pelajaran_list = MataPelajaran.objects.all().order_by('nama_pelajaran')
    kelas_list = Kelas.objects.all().order_by('nama')

    context = {
        'mata_pelajaran_list': mata_pelajaran_list,
        'kelas_list': kelas_list,
    }
    return render(request, 'admin/mata_pelajaran_admin.html', context)

@login_required
def get_mata_pelajaran_data(request, mapel_id):
    if request.method == 'GET' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        try:
            mapel = get_object_or_404(MataPelajaran, id=mapel_id)
            data = {
                'id': mapel.id,
                'nama_pelajaran': mapel.nama_pelajaran,
                'kode_pelajaran': mapel.kode_pelajaran,
                'kelas_id': mapel.kelas.id if mapel.kelas else None
            }
            return JsonResponse({'status': 'success', 'data': data})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=400)
    return JsonResponse({'status': 'error', 'message': 'Invalid request'}, status=400)

@csrf_exempt
def edit_siswa(request, pk):
    siswa = get_object_or_404(Siswa, pk=pk)
    if request.method == "POST":
        form = SiswaForm(request.POST, instance=siswa)
        new_password = request.POST.get('password')
        
        if form.is_valid():
            form.save()
            
            # Update password if provided
            if new_password and new_password.strip():
                if len(new_password) < 8:
                    messages.error(request, 'Password harus minimal 8 karakter!')
                    return render(request, 'admin/edit_siswa.html', {
                        'form': form,
                        'siswa': siswa,
                    })
                
                try:
                    siswa.user.set_password(new_password)
                    siswa.user.save()
                    messages.success(request, "Data siswa dan password berhasil diupdate.")
                except Exception as e:
                    messages.error(request, f'Gagal mengubah password: {str(e)}')
            else:
                messages.success(request, "Data siswa berhasil diupdate.")
            
            return redirect('rumahbelajar:data_siswa')
        else:
            messages.error(request, "Gagal mengupdate data siswa. Periksa inputan.")
    else:
        form = SiswaForm(instance=siswa)

    return render(request, 'admin/edit_siswa.html', {
        'form': form,
        'siswa': siswa,
    })

def edit_orang_tua(request, id):
    from rumahbelajar.models import OrangTua

    orang_tua = get_object_or_404(OrangTua, id=id)
    user = orang_tua.user  # akses user yang terhubung

    if request.method == 'POST':
        orang_tua.nama = request.POST.get('nama')
        orang_tua.no_telp = request.POST.get('no_telp')
        orang_tua.alamat = request.POST.get('alamat')
        
        # Update user.username
        username = request.POST.get('username')
        if username:
            orang_tua.user.username = username
            orang_tua.user.save()

        # Update password (jika diisi)
        password = request.POST.get('password')
        if password:
            orang_tua.user.set_password(password)
            orang_tua.user.save()

        orang_tua.save()
        return redirect('rumahbelajar:data_orang_tua')

    return render(request, 'admin/edit_orang_tua.html', {
        'orang_tua': orang_tua,
        'user': orang_tua.user
    })

def get_guru_data(request, id):
    guru = get_object_or_404(Guru, id=id)
    data = {
        'id': guru.id,
        'nama': guru.nama,
        'nip': guru.nip,
        'alamat': guru.alamat,
    }
    return JsonResponse(data)

def edit_guru(request):
    if request.method == 'POST':
        guru_id = request.POST.get('edit_id')
        nama = request.POST.get('edit_nama')
        nip = request.POST.get('edit_nip')
        alamat = request.POST.get('edit_alamat')
        new_password = request.POST.get('edit_password')

        guru = get_object_or_404(Guru, id=guru_id)
        guru.nama = nama
        guru.nip = nip
        guru.alamat_guru = alamat
        guru.save()

        # Update password if provided
        if new_password and new_password.strip():
            if len(new_password) < 8:
                messages.error(request, 'Password harus minimal 8 karakter!')
                return redirect('rumahbelajar:data_guru_admin')
            
            try:
                guru.user.set_password(new_password)
                guru.user.save()
                messages.success(request, 'Data guru dan password berhasil diperbarui.')
            except Exception as e:
                messages.error(request, f'Gagal mengubah password: {str(e)}')
        else:
            messages.success(request, 'Data guru berhasil diperbarui.')

        return redirect('rumahbelajar:data_guru_admin')
    else:
        messages.error(request, 'Permintaan tidak valid.')
        return redirect('rumahbelajar:data_guru_admin')


@login_required
@group_required('Owner')
def owner_siswa(request):
    siswa_list = Siswa.objects.all().order_by('nama')
    return render(request, 'owner/data_siswa_owner.html', {'siswa_list': siswa_list})

@login_required
@group_required('Owner')
def owner_orangtua(request):
    orangTua_list = OrangTua.objects.all().order_by('nama')
    return render(request, 'owner/data_orangtua_owner.html', {'orangTua_list': orangTua_list})

@login_required
@group_required('Owner')
def owner_guru(request):
    guru_list = Guru.objects.all().order_by('nama')
    return render(request, 'owner/data_guru_owner.html', {'guru_list': guru_list})

@login_required
@group_required('Owner')
def owner_mapel(request):
    mapel = MataPelajaran.objects.all().order_by('nama_pelajaran')
    return render(request, 'owner/data_mapel_owner.html', {'mapel': mapel})


@login_required
@user_passes_test(lambda u: u.groups.filter(name='Admin').exists())  # Cuma admin bisa akses
def generate_admin_qrcode(request):
    today = timezone.localdate()
    daily_qr, created = DailyQRCode.objects.get_or_create(tanggal=today)
    
    # Gunakan URL yang langsung redirect ke presensi guru
    url = request.build_absolute_uri(f"/absensi/{today}/")
    qr_img = qrcode.make(url)
    buffer = BytesIO()
    qr_img.save(buffer, format='PNG')
    daily_qr.qrcode.save(f'qr_{today}.png', ContentFile(buffer.getvalue()))
    daily_qr.save()
    
    return render(request, 'admin/admin_qrcode.html', {'daily_qr': daily_qr})

@login_required
@group_required('Siswa')
def scan_qr_absen(request, presensi_id):
    try:
        siswa = Siswa.objects.get(user=request.user)
        presensi = get_object_or_404(Presensi, id=presensi_id)

        # Cek apakah presensi sesuai dengan kelas siswa
        if presensi.kelas != siswa.kelas:
            messages.error(request, "QR code tidak sesuai dengan kelas Anda.")
            return redirect('rumahbelajar:absensi_siswa')

        # Cek apakah sudah absen
        if PresensiSiswa.objects.filter(presensi=presensi, siswa=siswa).exists():
            messages.info(request, "Anda sudah melakukan absensi.")
        else:
            PresensiSiswa.objects.create(presensi=presensi, siswa=siswa, status='hadir')
            messages.success(request, "Absensi berhasil direkam.")

        return redirect('rumahbelajar:absensi_siswa')
    except Siswa.DoesNotExist:
        messages.error(request, 'Akun siswa tidak valid.')
        return redirect('rumahbelajar:login')

    
@login_required
@group_required('Admin')  # atau guru
def generate_qr_presensi(request):
    presensi_list = Presensi.objects.all().order_by('-id')
    qr_data = []

    for presensi in presensi_list:
        url = request.build_absolute_uri(f"/absensi/scan/{presensi.id}/")
        qr = qrcode.make(url)
        buffer = BytesIO()
        qr.save(buffer)
        img_base64 = base64.b64encode(buffer.getvalue()).decode()
        qr_data.append({'presensi': presensi, 'qr_image': img_base64})

    return render(request, 'admin/admin_siswa_qrcode.html', {'qr_data': qr_data})

@login_required
def akademik_anak_view(request):
    orangtua = get_object_or_404(OrangTua, user=request.user)
    siswa_list = Siswa.objects.filter(orang_tua=orangtua)

    if not siswa_list.exists():
        context = {
            'error': 'Data siswa tidak ditemukan untuk akun ini.'
        }
        return render(request, 'orangtua/akademik_anak.html', context)

    # Ambil semua mata pelajaran per anak berdasarkan kelas masing-masing
    siswa_data = []
    for siswa in siswa_list:
        kelas = siswa.kelas
        mata_pelajaran_list = MataPelajaran.objects.filter(kelas=kelas) if kelas else []
        siswa_data.append({
            'siswa': siswa,
            'kelas': kelas,
            'mata_pelajaran_list': mata_pelajaran_list
        })

    context = {
        'siswa_data': siswa_data,
    }

    return render(request, 'orangtua/akademik_anak.html', context)

@login_required
@group_required('Siswa')
def daftar_mapel_siswa_view(request):
    try:
        siswa = Siswa.objects.get(user=request.user)
        # Ambil mata pelajaran sesuai kelas siswa
        mata_pelajaran_list = MataPelajaran.objects.filter(kelas=siswa.kelas)
        
        context = {
            'siswa': siswa,
            'mata_pelajaran_list': mata_pelajaran_list,
        }
        return render(request, 'siswa/daftar_mapel.html', context)
    except Siswa.DoesNotExist:
        messages.error(request, 'Akun siswa tidak valid.')
        return redirect('rumahbelajar:login')



        return redirect('rumahbelajar:absensi_detail', jadwal_id=jadwal_id)

import qrcode
from io import BytesIO
from django.core.files.base import ContentFile


@login_required
def generate_qr_guru(request, jadwal_id, pertemuan):
    kode = f"{jadwal_id}-{pertemuan}"

    qr_obj, created = QRCode.objects.get_or_create(
        jadwal_id=jadwal_id,
        pertemuan=pertemuan,
        defaults={'kode': kode}
    )

    if created or not qr_obj.gambar_qr:
        img = qrcode.make(kode)
        buffer = BytesIO()
        img.save(buffer, format='PNG')
        file_name = f'qr_{jadwal_id}_{pertemuan}.png'
        qr_obj.gambar_qr.save(file_name, ContentFile(buffer.getvalue()), save=True)
        print("QR disimpan:", file_name)
    context = {
        'kode': kode,
        'qr_obj': qr_obj,
        'jadwal_id': jadwal_id,
        'pertemuan': pertemuan,
    }
    return render(request, 'guru/generate_qr.html', context) # ganti dengan nama halaman kamu


def pilih_anak_view(request):
    orangtua = request.user.rumahbelajar_orangtua # asumsinya User punya relasi OneToOne dengan OrangTua
    siswa_list = Siswa.objects.filter(orang_tua=orangtua)
    return render(request, 'orangtua/pilih_anak.html', {'siswa_list': siswa_list})


def daftar_mapel_anak_view(request, siswa_id):
    siswa = get_object_or_404(Siswa, id=siswa_id, orang_tua=request.user.rumahbelajar_orangtua)
    kelas = siswa.kelas
    mata_pelajaran_list = MataPelajaran.objects.filter(kelas=kelas)
    return render(request, 'orangtua/absensi_daftar_mapel.html', {
        'siswa': siswa,
        'mata_pelajaran_list': mata_pelajaran_list
    })


@login_required
def absensi_detail_anak(request, presensi_id, siswa_id):
    orang_tua = get_object_or_404(OrangTua, user=request.user)
    siswa = get_object_or_404(Siswa, id=siswa_id, orang_tua=orang_tua)
    jadwal = get_object_or_404(Presensi, id=presensi_id)

    total_pertemuan = range(1, 17)
    presensi_siswa = PresensiSiswa.objects.filter(presensi=jadwal, siswa=siswa)

    status_pertemuan = {p.pertemuan_ke: p.status for p in presensi_siswa}
    presensi_waktu = {p.pertemuan_ke: p.waktu_presensi for p in presensi_siswa}

    context = {
        'jadwal': jadwal,
        'siswa': siswa,
        'total_pertemuan': total_pertemuan,
        'status_pertemuan': status_pertemuan,
        'presensi_waktu': presensi_waktu,
        'year': timezone.now().year,
    }
    return render(request, 'orangtua/absensi_detail.html', context)



from django.shortcuts import render, redirect
from .forms import PresensiForm

def tambah_presensi(request):
    if request.method == 'POST':
        form = PresensiForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('rumahbelajar:dashboard_admin')  # ganti jika perlu
    else:
        form = PresensiForm()

    return render(request, 'admin/tambah_presensi.html', {'form': form})

@login_required
def change_password(request):
    """View untuk mengubah password user"""
    if request.method == 'POST':
        current_password = request.POST.get('current_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        # Validasi password saat ini
        if not request.user.check_password(current_password):
            messages.error(request, 'Password saat ini tidak benar!')
            return redirect(request.META.get('HTTP_REFERER', '/'))
        
        # Validasi password baru
        if new_password != confirm_password:
            messages.error(request, 'Password baru dan konfirmasi password tidak cocok!')
            return redirect(request.META.get('HTTP_REFERER', '/'))
        
        if len(new_password) < 8:
            messages.error(request, 'Password baru harus minimal 8 karakter!')
            return redirect(request.META.get('HTTP_REFERER', '/'))
        
        # Update password
        try:
            request.user.set_password(new_password)
            request.user.save()
            update_session_auth_hash(request, request.user)  # Keep user logged in
            messages.success(request, 'Password berhasil diubah!')
        except Exception as e:
            messages.error(request, f'Terjadi kesalahan: {str(e)}')
        
        return redirect(request.META.get('HTTP_REFERER', '/'))
    
    return redirect(request.META.get('HTTP_REFERER', '/'))

@login_required
@group_required('Siswa')
def riwayat_presensi_siswa(request):
    """View untuk menampilkan riwayat presensi siswa dengan filter tanggal dan tahun"""
    try:
        siswa = Siswa.objects.get(user=request.user)
        
        # Filter by date range and year
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        selected_year = request.GET.get('year', timezone.now().year)
        print(selected_year)
        year_range = get_year_range_for_model(PresensiSiswa, 'waktu_presensi')
        
        presensi_list = PresensiSiswa.objects.filter(siswa=siswa).select_related('presensi', 'presensi__mata_pelajaran', 'presensi__kelas')
        
        if selected_year:
            presensi_list = presensi_list.filter(waktu_presensi__year=selected_year)
        
        if start_date:
            presensi_list = presensi_list.filter(waktu_presensi__date__gte=start_date)
        if end_date:
            presensi_list = presensi_list.filter(waktu_presensi__date__lte=end_date)
        
        presensi_list = presensi_list.order_by('-waktu_presensi')
        
        return render(request, 'siswa/riwayat_presensi.html', {
            'siswa': siswa,
            'presensi_list': presensi_list,
            'start_date': start_date,
            'end_date': end_date,
            'selected_year': selected_year,
            'year_range': year_range,
        })
    except Siswa.DoesNotExist:
        messages.error(request, 'Akun siswa tidak valid.')
        return redirect('rumahbelajar:login')

@login_required
def riwayat_presensi_anak(request, siswa_id):
    """View untuk orangtua melihat riwayat presensi anak dengan filter tanggal dan tahun"""
    try:
        orangtua = OrangTua.objects.get(user=request.user)
        siswa = get_object_or_404(Siswa, id=siswa_id, orang_tua=orangtua)
    except OrangTua.DoesNotExist:
        messages.error(request, 'Data orangtua tidak ditemukan!')
        return redirect('rumahbelajar:dashboard_OrangTua')
    
    # Filter by date range and year
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    selected_year = request.GET.get('year')
    
    # Get year range for dropdown
    year_range = get_year_range_for_model(PresensiSiswa, 'waktu_presensi')
    
    # Get attendance records
    presensi_list = PresensiSiswa.objects.filter(siswa=siswa).select_related('presensi', 'presensi__mata_pelajaran', 'presensi__kelas')
    
    # Apply year filter if provided
    if selected_year:
        presensi_list = presensi_list.filter(waktu_presensi__year=selected_year)
    
    # Apply date filter if provided
    if start_date:
        presensi_list = presensi_list.filter(waktu_presensi__date__gte=start_date)
    if end_date:
        presensi_list = presensi_list.filter(waktu_presensi__date__lte=end_date)
    
    # Order by date (newest first)
    presensi_list = presensi_list.order_by('-waktu_presensi')
    
    return render(request, 'orangtua/riwayat_presensi_anak.html', {
        'siswa': siswa,
        'presensi_list': presensi_list,
        'start_date': start_date,
        'end_date': end_date,
        'selected_year': selected_year,
        'year_range': year_range,
    })

@login_required
def check_teacher_attendance(request):
    """AJAX endpoint to check if teacher has marked attendance today"""
    try:
        guru = Guru.objects.get(user=request.user)
        print(guru)
        today = timezone.now().date()
        guru_has_attended = PresensiGuru.objects.filter(
            guru=guru,
            tanggal=today,
            status='Hadir'
        ).exists()
        
        return JsonResponse({
            'has_attended': guru_has_attended,
            'today': today.isoformat(),
            'guru_name': guru.nama
        })
    except Guru.DoesNotExist:
        return JsonResponse({
            'error': 'Data guru tidak ditemukan',
            'has_attended': False
        }, status=404)

@login_required
def absensi_guru_daily(request, tanggal):
    print(f"DEBUG: absensi_guru_daily dipanggil dengan tanggal: {tanggal}")
    print(f"DEBUG: User: {request.user.username}")
    
    # Validasi user harus guru
    try:
        guru_obj = Guru.objects.get(user=request.user)
        print(f"DEBUG: Guru ditemukan: {guru_obj.nama}")
    except Guru.DoesNotExist:
        print("DEBUG: User bukan guru")
        messages.error(request, 'Hanya guru yang bisa absen lewat QR code.')
        return redirect('rumahbelajar:scan_qr_view')

    # Validasi format tanggal
    try:
        tanggal_qr = datetime.datetime.strptime(tanggal, "%Y-%m-%d").date()
        print(f"DEBUG: Tanggal QR: {tanggal_qr}")
    except ValueError:
        print("DEBUG: Format tanggal tidak valid")
        messages.error(request, 'Format tanggal tidak valid.')
        return redirect('rumahbelajar:scan_qr_view')

    # Validasi QR hanya untuk hari ini
    tanggal_hari_ini = localdate()
    print(f"DEBUG: Tanggal hari ini: {tanggal_hari_ini}")
    if tanggal_qr != tanggal_hari_ini:
        print("DEBUG: QR bukan untuk hari ini")
        messages.error(request, 'QR code ini bukan untuk hari ini.')
        return redirect('rumahbelajar:scan_qr_view')

    # Simpan absensi hari ini ke database (model Absensi)
    print("DEBUG: Mencoba menyimpan absensi...")
    absen, created = Absensi.objects.get_or_create(
        guru=guru_obj,
        tanggal=tanggal_hari_ini,
        defaults={
            'jam_absensi': timezone.localtime().time(),
            'status': 'Hadir',
            'metode_absensi': 'Absen via QR'
        }
    )
    
    if created:
        print(f"DEBUG: Absensi baru dibuat - ID: {absen.id}")
    else:
        print(f"DEBUG: Absensi sudah ada - ID: {absen.id}, update data...")
        absen.status = 'Hadir'
        absen.metode_absensi = 'Absen via QR'
        absen.jam_absensi = timezone.localtime().time()
        absen.save()
        print(f"DEBUG: Absensi diupdate - Status: {absen.status}, Metode: {absen.metode_absensi}")

    # Simpan presensi guru ke database (model PresensiGuru)
    print("DEBUG: Mencoba menyimpan presensi guru...")
    presensi_guru, presensi_created = PresensiGuru.objects.get_or_create(
        guru=guru_obj,
        tanggal=tanggal_hari_ini,
        defaults={
            'jam_masuk': timezone.localtime().time(),
            'status': 'Hadir',
            'keterangan': 'Absen via QR Code'
        }
    )
    
    if presensi_created:
        print(f"DEBUG: PresensiGuru baru dibuat - ID: {presensi_guru.id}")
        messages.success(request, 'Absensi dan presensi guru berhasil dicatat!')
    else:
        print(f"DEBUG: PresensiGuru sudah ada - ID: {presensi_guru.id}, update data...")
        presensi_guru.status = 'Hadir'
        presensi_guru.jam_masuk = timezone.localtime().time()
        presensi_guru.keterangan = 'Absen via QR Code'
        presensi_guru.save()
        print(f"DEBUG: PresensiGuru diupdate - Status: {presensi_guru.status}")
        messages.success(request, 'Absensi dan presensi guru berhasil diperbarui!')

    # Redirect ke halaman presensi guru
    return redirect('rumahbelajar:presensi_guru')
